import streamlit as st
import pandas as pd
import os
import subprocess
from datetime import datetime, timedelta
from io import BytesIO
from supabase import create_client, Client

# ============================================================
# CONFIGURACIÓN DE PÁGINA
# ============================================================
st.set_page_config(
    page_title="Concierge Master",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ============================================================
# CSS PERSONALIZADO - TEMA OSCURO
# ============================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

* {
    font-family: 'Inter', sans-serif !important;
}

/* Fondo general */
.stApp {
    background-color: #0a0a0a !important;
    color: #e0e0e0 !important;
}

/* Ocultar header default y ajustar layout */
header[data-testid="stHeader"] { display: none !important; }
.block-container { 
    padding-top: 0.5rem !important; 
    padding-bottom: 0.5rem !important;
    max-width: 100% !important;
}

/* Scrollbar personalizada */
::-webkit-scrollbar {
    width: 8px;
    height: 8px;
}
::-webkit-scrollbar-track {
    background: #1a1a1a;
}
::-webkit-scrollbar-thumb {
    background: #333;
    border-radius: 4px;
}
::-webkit-scrollbar-thumb:hover {
    background: #555;
}

/* === HEADER === */
.header-container {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 0.5rem 1rem;
    background: linear-gradient(90deg, #0a0a0a 0%, #111 50%, #0a0a0a 100%);
    border-bottom: 1px solid #222;
    margin-bottom: 0.5rem;
}
.header-left {
    display: flex;
    align-items: center;
    gap: 12px;
}
.header-logo {
    font-size: 1.2rem;
    font-weight: 700;
    color: #d4af37;
    letter-spacing: 2px;
}
.header-sub {
    font-size: 0.75rem;
    color: #00E5FF;
    font-weight: 600;
    letter-spacing: 1px;
}
.header-right {
    text-align: right;
}
.header-date {
    font-size: 0.85rem;
    color: #00E5FF;
    font-weight: 600;
}
.header-time {
    font-size: 1.1rem;
    color: #00E5FF;
    font-weight: 700;
    letter-spacing: 1px;
}

/* === BARRA TOTAL === */
.total-bar {
    background: linear-gradient(90deg, #0d1b2a 0%, #1b2838 50%, #0d1b2a 100%);
    border: 1px solid #00E5FF;
    border-radius: 8px;
    padding: 0.6rem 1rem;
    text-align: center;
    margin-bottom: 0.8rem;
    box-shadow: 0 0 15px rgba(0, 229, 255, 0.1);
}
.total-bar span {
    color: #00E5FF;
    font-size: 0.9rem;
    font-weight: 700;
    letter-spacing: 1px;
}

/* === PANEL CHECKING OUT === */
.checkout-panel {
    background: #111;
    border: 1px solid #222;
    border-radius: 8px;
    padding: 0.8rem;
    margin-bottom: 0.8rem;
}
.checkout-title {
    color: #888;
    font-size: 0.65rem;
    font-weight: 700;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    margin-bottom: 0.5rem;
    border-bottom: 1px solid #222;
    padding-bottom: 0.3rem;
}
.checkout-item {
    display: flex;
    justify-content: space-between;
    align-items: center;
    background: #1a1a1a;
    border-radius: 4px;
    padding: 0.4rem 0.6rem;
    margin-bottom: 0.3rem;
    border: 1px solid #2a2a2a;
}
.checkout-date {
    font-size: 0.7rem;
    font-weight: 700;
    color: #ccc;
}
.checkout-count {
    font-size: 0.75rem;
    font-weight: 700;
    color: #fff;
    background: #333;
    border-radius: 50%;
    width: 20px;
    height: 20px;
    display: flex;
    align-items: center;
    justify-content: center;
}
.btn-ver-todas {
    width: 100%;
    background: #1a1a1a !important;
    color: #888 !important;
    border: 1px solid #333 !important;
    border-radius: 4px !important;
    font-size: 0.65rem !important;
    padding: 0.3rem !important;
    margin-top: 0.3rem !important;
}

/* === CATEGORÍAS === */
.cat-container {
    background: #111;
    border: 1px solid #222;
    border-radius: 8px;
    padding: 0.8rem;
    margin-bottom: 0.8rem;
}
.cat-title {
    color: #888;
    font-size: 0.65rem;
    font-weight: 700;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    margin-bottom: 0.6rem;
}
.cat-row {
    display: flex;
    align-items: center;
    margin-bottom: 0.4rem;
    gap: 8px;
}
.cat-label {
    width: 90px;
    font-size: 0.65rem;
    color: #aaa;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    text-align: right;
}
.cat-bar-bg {
    flex: 1;
    height: 14px;
    background: #1a1a1a;
    border-radius: 7px;
    overflow: hidden;
    position: relative;
}
.cat-bar-fill {
    height: 100%;
    border-radius: 7px;
    transition: width 0.5s ease;
}
.cat-count {
    width: 20px;
    text-align: right;
    font-size: 0.75rem;
    font-weight: 700;
    color: #fff;
}

/* Colores de categorías */
.cat-vip { background: linear-gradient(90deg, #FFD700, #FFA500); }
.cat-birthday { background: linear-gradient(90deg, #FF6B6B, #FF8E8E); }
.cat-honeymoon { background: linear-gradient(90deg, #FF69B4, #FF1493); }
.cat-anniversary { background: linear-gradient(90deg, #4CAF50, #8BC34A); }
.cat-babymoon { background: linear-gradient(90deg, #9C27B0, #E1BEE7); }
.cat-team { background: linear-gradient(90deg, #FF9800, #FFC107); }
.cat-leisure { background: linear-gradient(90deg, #2196F3, #03A9F4); }
.cat-relaxury { 
    background: linear-gradient(90deg, #E91E63, #F48FB1);
    box-shadow: 0 0 10px rgba(233, 30, 99, 0.3);
}

/* === BOTONES DE ACCIÓN === */
.action-btn {
    border: none !important;
    border-radius: 4px !important;
    font-size: 0.65rem !important;
    font-weight: 700 !important;
    padding: 0.4rem 0.8rem !important;
    color: #000 !important;
    cursor: pointer;
    letter-spacing: 0.5px;
    text-transform: uppercase;
}
.btn-nueva { background: #00E5FF !important; }
.btn-import { background: #4CAF50 !important; color: #fff !important; }
.btn-export { background: #2196F3 !important; color: #fff !important; }
.btn-reporte { background: #FFC107 !important; }
.btn-bonus { background: #FF9800 !important; }
.btn-qiaf { background: #FFD700 !important; }

/* Botones de restaurantes */
.rest-btn {
    border: none !important;
    border-radius: 4px !important;
    font-size: 0.6rem !important;
    font-weight: 600 !important;
    padding: 0.3rem 0.6rem !important;
    color: #fff !important;
    cursor: pointer;
}
.rest-alice { background: #5C6BC0 !important; }
.rest-arrivals { background: #EF5350 !important; }
.rest-lacena { background: #66BB6A !important; }
.rest-nolimit { background: #AB47BC !important; }
.rest-opentable { background: #EC407A !important; }
.rest-outlookfw { background: #42A5F5 !important; }
.rest-outlookpc { background: #7E57C2 !important; }
.rest-relaxury { background: #78909C !important; }
.rest-vtc { background: #8D6E63 !important; }

/* === FILTROS === */
.filter-panel {
    background: #111;
    border: 1px solid #222;
    border-radius: 8px;
    padding: 0.8rem;
    margin-bottom: 0.8rem;
}
.filter-label {
    color: #888;
    font-size: 0.6rem;
    font-weight: 700;
    letter-spacing: 1px;
    text-transform: uppercase;
    margin-bottom: 0.3rem;
}
.filter-date {
    background: #1a1a1a !important;
    color: #fff !important;
    border: 1px solid #333 !important;
    border-radius: 4px !important;
}

/* === BARRA DE BÚSQUEDA === */
.search-bar {
    background: #111;
    border: 1px solid #222;
    border-radius: 8px;
    padding: 0.6rem 1rem;
    margin-bottom: 0.8rem;
    display: flex;
    align-items: center;
    gap: 10px;
}
.search-icon {
    color: #666;
    font-size: 1rem;
}
.search-input {
    background: transparent !important;
    border: none !important;
    color: #fff !important;
    font-size: 0.85rem !important;
    width: 100%;
}
.search-input::placeholder {
    color: #555 !important;
}

/* === TABLA === */
.table-container {
    background: #111;
    border: 1px solid #222;
    border-radius: 8px;
    overflow: hidden;
}
.table-header {
    background: #1a1a1a;
    color: #00E5FF;
    font-size: 0.65rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1px;
    padding: 0.6rem;
    border-bottom: 2px solid #00E5FF;
}
.table-row {
    border-bottom: 1px solid #1a1a1a;
    font-size: 0.75rem;
    color: #ccc;
}
.table-row:hover {
    background: #1a1a1a !important;
}

/* Colores condicionales en tabla */
.text-vip { color: #FFD700 !important; font-weight: 700 !important; }
.text-birthday { color: #FF6B6B !important; font-weight: 700 !important; }
.text-honeymoon { color: #FF69B4 !important; font-weight: 700 !important; }
.text-team { color: #FF9800 !important; font-weight: 700 !important; }
.text-relaxury { color: #E91E63 !important; font-weight: 700 !important; }
.text-gold { color: #FFD700 !important; font-weight: 700 !important; }
.text-silver { color: #C0C0C0 !important; font-weight: 700 !important; }
.text-diamond { color: #B9F2FF !important; font-weight: 700 !important; }

/* === SPLASH SCREEN === */
#splash-overlay {
    position: fixed;
    top: 0;
    left: 0;
    width: 100vw;
    height: 100vh;
    background-color: #0a0a0a;
    z-index: 999999;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    animation: splashFade 4.5s ease-in-out forwards;
}
#splash-overlay img {
    width: 100%;
    height: 100%;
    object-fit: cover;
    object-position: center;
    position: absolute;
    top: 0;
    left: 0;
    opacity: 0.3;
}
#splash-content {
    position: relative;
    z-index: 2;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: flex-end;
    width: 100%;
    height: 100%;
    padding-bottom: 80px;
}
#progress-container {
    width: 320px;
    height: 3px;
    background-color: rgba(212, 175, 55, 0.15);
    border-radius: 3px;
    overflow: hidden;
    margin-top: 20px;
}
#progress-bar {
    width: 0%;
    height: 100%;
    background: linear-gradient(90deg, #d4af37, #f0d878, #d4af37);
    border-radius: 3px;
    animation: progressFill 3.5s ease-out forwards;
    box-shadow: 0 0 10px rgba(212, 175, 55, 0.5);
}
#loading-text {
    color: #d4af37;
    font-family: 'Segoe UI', sans-serif;
    font-size: 0.7rem;
    letter-spacing: 4px;
    text-transform: uppercase;
    margin-top: 15px;
    opacity: 0.8;
    animation: textPulse 1.5s ease-in-out infinite;
}
@keyframes progressFill {
    0% { width: 0%; }
    20% { width: 15%; }
    50% { width: 60%; }
    80% { width: 85%; }
    100% { width: 100%; }
}
@keyframes textPulse {
    0%, 100% { opacity: 0.5; }
    50% { opacity: 1; }
}
@keyframes splashFade {
    0% { opacity: 1; }
    78% { opacity: 1; }
    100% { opacity: 0; visibility: hidden; }
}

/* === FORMULARIOS === */
.stForm {
    background: #111 !important;
    border: 1px solid #222 !important;
    border-radius: 8px !important;
    padding: 1rem !important;
}
.stTextInput > div > div > input {
    background: #1a1a1a !important;
    color: #fff !important;
    border: 1px solid #333 !important;
    border-radius: 4px !important;
}
.stSelectbox > div > div > div {
    background: #1a1a1a !important;
    color: #fff !important;
    border: 1px solid #333 !important;
}
.stDateInput > div > div > input {
    background: #1a1a1a !important;
    color: #fff !important;
    border: 1px solid #333 !important;
}
.stTextArea > div > div > textarea {
    background: #1a1a1a !important;
    color: #fff !important;
    border: 1px solid #333 !important;
}

/* === MISC === */
.divider {
    height: 1px;
    background: #222;
    margin: 0.5rem 0;
}
.info-badge {
    display: inline-block;
    padding: 2px 6px;
    border-radius: 3px;
    font-size: 0.65rem;
    font-weight: 700;
    margin-right: 4px;
}
</style>
""", unsafe_allow_html=True)

# ============================================================
# SPLASHSCREEN
# ============================================================
import time

if "splash_shown" not in st.session_state:
    st.session_state["splash_shown"] = True

    splash_html = """
    <div id="splash-overlay">
        <img src="data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMTAwIiBoZWlnaHQ9IjEwMCIgeG1sbnM9Imh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIj48cmVjdCB3aWR0aD0iMTAwJSIgaGVpZ2h0PSIxMDAlIiBmaWxsPSIjMGEwYTBhIi8+PC9zdmc+" alt="Concierge Master" />
        <div id="splash-content">
            <div style="text-align: center; margin-bottom: 40px;">
                <div style="font-size: 2.5rem; font-weight: 700; color: #d4af37; letter-spacing: 4px; margin-bottom: 8px;">CONCIERGE</div>
                <div style="font-size: 1rem; color: #00E5FF; letter-spacing: 8px; font-weight: 300;">MASTER</div>
                <div style="font-size: 0.7rem; color: #666; margin-top: 8px; letter-spacing: 2px;">v5.1</div>
            </div>
            <div id="progress-container">
                <div id="progress-bar"></div>
            </div>
            <div id="loading-text">Loading Concierge Master</div>
        </div>
    </div>
    """
    st.html(splash_html)
    time.sleep(0.1)

# ============================================================
# CONEXIÓN SUPABASE
# ============================================================
@st.cache_resource
def init_supabase():
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

supabase = init_supabase()
TABLE_NAME = "huespedes"

# ============================================================
# PARÁMETROS DE URL
# ============================================================
query_params = st.query_params
mostrar_formulario = query_params.get("action") == "nueva"
mostrar_editar = query_params.get("action") == "editar"
mostrar_importar = query_params.get("action") == "importar"
mostrar_exportar = query_params.get("action") == "exportar"
filtro_checkout = query_params.get("checkout_filtro")
filtro_fecha_date = query_params.get("fecha_date")
fecha_filtro_activo = query_params.get("fecha_activa") == "true"

# ============================================================
# FUNCIONES CRUD
# ============================================================
@st.cache_data(ttl=30)
def cargar_reservaciones():
    response = supabase.table(TABLE_NAME).select("*").execute()
    df = pd.DataFrame(response.data)
    if df.empty:
        return pd.DataFrame(columns=["id", "eta", "name", "qty", "room", "email",
                                     "check_in", "check_out", "res_number", "phone",
                                     "info", "ird", "hsk", "rate", "trans"])
    df["check_in_dt"] = pd.to_datetime(df["check_in"], errors="coerce")
    df = df.sort_values(by=["check_in_dt", "name"])
    return df.drop(columns=["check_in_dt"])

def insertar_reserva(data: dict):
    supabase.table(TABLE_NAME).insert(data).execute()
    st.cache_data.clear()

def actualizar_reserva(reserva_id, data: dict):
    supabase.table(TABLE_NAME).update(data).eq("id", reserva_id).execute()
    st.cache_data.clear()

def eliminar_reserva(reserva_id):
    supabase.table(TABLE_NAME).delete().eq("id", reserva_id).execute()
    st.cache_data.clear()

def insertar_batch_reservas(lista_data: list):
    supabase.table(TABLE_NAME).insert(lista_data).execute()
    st.cache_data.clear()

# ============================================================
# HELPERS
# ============================================================
def parse_fecha(fecha_str):
    if not fecha_str or str(fecha_str).strip() == "":
        return None
    s = str(fecha_str).strip()
    for fmt in ["%B %d, %Y", "%b %d, %Y", "%B %d", "%b %d", "%Y-%m-%d"]:
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return pd.to_datetime(s, errors="coerce")

def duracion_noches(check_in, check_out):
    try:
        ci = pd.to_datetime(check_in)
        co = pd.to_datetime(check_out)
        return (co - ci).days
    except:
        return ""

def hora_actual_12h():
    ahora = datetime.now()
    h, m = ahora.hour, ahora.minute
    if m >= 45:
        m = 0
        h = (h + 1) % 24
    elif m >= 15:
        m = 30
    else:
        m = 0
    if h == 0:
        return f"12:{m:02d} AM"
    elif h < 12:
        return f"{h}:{m:02d} AM"
    elif h == 12:
        return f"12:{m:02d} PM"
    else:
        return f"{h-12}:{m:02d} PM"

# ============================================================
# EXPORTAR EXCEL
# ============================================================
def exportar_excel_por_categorias(df):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    fill_data = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_header = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    fill_col_header = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10, color="000000")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Arrivals"
    
    categorias_config = [
        ("CUMPLEAÑOS", ["BIRTHDAY", "CUMPLE", "BDAY"]),
        ("VIP", ["VIP"]),
        ("HONEYMOON", ["HONEYMOON", "LUNA DE MIEL"]),
        ("ANNIVERSARY", ["ANNIVERSARY", "ANIVERSARIO"]),
        ("BABYMOON", ["BABYMOON"]),
        ("TEAM MEMBER", ["TEAM MEMBER", "STAFF", "EMPLOYEE"]),
        ("GENERAL", [])
    ]
    
    columnas_excel = ["ID", "ETA", "NAME", "QTY", "ROOM", "EMAIL", "CHECK IN", "CHECK OUT",
                      "RESERVATION", "PHONE", "FORMATIO", "IRD", "HSK", "RATE", "TRANSPORTATION"]
    mapeo_cols = {
        "id": "ID", "eta": "ETA", "name": "NAME", "qty": "QTY", "room": "ROOM",
        "email": "EMAIL", "check_in": "CHECK IN", "check_out": "CHECK OUT",
        "res_number": "RESERVATION", "phone": "PHONE", "info": "FORMATIO",
        "ird": "IRD", "hsk": "HSK", "rate": "RATE", "trans": "TRANSPORTATION"
    }
    
    current_row = 1
    filas_por_categoria, filas_general = {}, []
    
    for _, row in df.iterrows():
        info_str = str(row.get("info", "")).upper()
        asignada = False
        for cat_nombre, keywords in categorias_config[:-1]:
            for kw in keywords:
                if kw in info_str:
                    filas_por_categoria.setdefault(cat_nombre, []).append(row)
                    asignada = True
                    break
            if asignada:
                break
        if not asignada:
            filas_general.append(row)
    
    for cat_nombre, keywords in categorias_config:
        filas = filas_general if cat_nombre == "GENERAL" else filas_por_categoria.get(cat_nombre, [])
        if not filas:
            continue
        
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columnas_excel))
        cell = ws.cell(row=current_row, column=1, value=cat_nombre)
        cell.fill = fill_header
        cell.font = font_section
        cell.alignment = align_center
        current_row += 1
        
        for col_idx, col_name in enumerate(columnas_excel, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.fill = fill_col_header
            cell.font = font_header
            cell.alignment = align_center
            cell.border = thin_border
        current_row += 1
        
        for row_data in filas:
            for col_idx, col_db in enumerate(mapeo_cols.keys(), 1):
                valor = row_data.get(col_db, "")
                if pd.isna(valor):
                    valor = ""
                cell = ws.cell(row=current_row, column=col_idx, value=valor)
                cell.fill = fill_data
                cell.font = font_data
                cell.alignment = align_left
                cell.border = thin_border
            current_row += 1
        current_row += 1
    
    anchos = {
        "A": 6, "B": 10, "C": 22, "D": 6, "E": 8, "F": 25, "G": 10, "H": 10,
        "I": 14, "J": 18, "K": 22, "L": 18, "M": 18, "N": 8, "O": 18
    }
    for col_letter, ancho in anchos.items():
        ws.column_dimensions[col_letter].width = ancho
    
    ws.freeze_panes = "A1"
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ============================================================
# HEADER PRINCIPAL
# ============================================================
def render_header():
    now = datetime.now()
    fecha_str = now.strftime("%B %d, %Y")
    hora_str = now.strftime("%I:%M:%S %p").lstrip("0")
    
    header_html = f"""
    <div class="header-container">
        <div class="header-left">
            <div>
                <div class="header-logo">WALDORF ASTORIA</div>
                <div style="font-size: 0.6rem; color: #666; letter-spacing: 1px;">COSTA RICA • PUNTA CANCUN</div>
            </div>
            <div style="width: 1px; height: 30px; background: #333; margin: 0 8px;"></div>
            <div>
                <div class="header-sub">Concierge</div>
                <div style="font-size: 0.65rem; color: #888;">Master v5.1</div>
            </div>
        </div>
        <div class="header-right">
            <div class="header-date">{fecha_str}</div>
            <div class="header-time">{hora_str}</div>
        </div>
    </div>
    """
    st.markdown(header_html, unsafe_allow_html=True)

# ============================================================
# BARRA DE CATEGORÍAS
# ============================================================
def render_categorias(df):
    cats = {
        "VIP": ("cat-vip", 0),
        "Birthday": ("cat-birthday", 0),
        "Honeymoon": ("cat-honeymoon", 0),
        "Anniversary": ("cat-anniversary", 0),
        "Babymoon": ("cat-babymoon", 0),
        "Team Member": ("cat-team", 0),
        "Leisure": ("cat-leisure", 0),
    }
    
    for _, row in df.iterrows():
        info = str(row.get("info", "")).upper()
        for key in cats:
            if key.upper() in info:
                cats[key] = (cats[key][0], cats[key][1] + 1)
    
    # Contar RelaXury por separado
    relaxury_count = df[df["trans"].astype(str).str.upper().str.contains("RELAXURY", na=False)].shape[0]
    
    max_val = max([v[1] for v in cats.values()]) if cats else 1
    if max_val == 0:
        max_val = 1
    
    html = '<div class="cat-container"><div class="cat-title">Guest Categories</div>'
    
    for label, (css_class, count) in cats.items():
        pct = (count / max_val) * 100 if max_val > 0 else 0
        html += f"""
        <div class="cat-row">
            <div class="cat-label">{label}</div>
            <div class="cat-bar-bg">
                <div class="cat-bar-fill {css_class}" style="width: {pct}%;"></div>
            </div>
            <div class="cat-count">{count}</div>
        </div>
        """
    
    # RelaXury bar
    html += f"""
    <div style="margin-top: 0.5rem; padding-top: 0.5rem; border-top: 1px solid #222;">
        <div class="cat-row">
            <div class="cat-label" style="color: #E91E63;">♥ RELAXURY</div>
            <div class="cat-bar-bg">
                <div class="cat-bar-fill cat-relaxury" style="width: 100%;"></div>
            </div>
            <div class="cat-count" style="color: #E91E63;">{relaxury_count}</div>
        </div>
    </div>
    """
    
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)

# ============================================================
# PANEL CHECKING OUT
# ============================================================
def render_checking_out(df):
    st.markdown('<div class="checkout-panel">', unsafe_allow_html=True)
    st.markdown('<div class="checkout-title">📋 Checking Out</div>', unsafe_allow_html=True)
    
    hoy = datetime.now().date()
    fechas_checkout = {}
    
    for _, row in df.iterrows():
        try:
            co = pd.to_datetime(row["check_out"]).date()
            if co >= hoy:
                key = co.strftime("%d-%b").upper()
                fechas_checkout[key] = fechas_checkout.get(key, 0) + 1
        except:
            pass
    
    fechas_ordenadas = sorted(fechas_checkout.items(), key=lambda x: datetime.strptime(x[0] + "-2026", "%d-%b-%Y"))
    
    for fecha, count in fechas_ordenadas[:7]:
        st.markdown(f"""
        <div class="checkout-item">
            <div class="checkout-date">{fecha}</div>
            <div class="checkout-count">{count}</div>
        </div>
        """, unsafe_allow_html=True)
    
    if st.button("🔄 VER TODAS", key="btn_ver_todas", use_container_width=True):
        st.query_params["action"] = "checkout_todas"
        st.rerun()
    
    st.markdown('</div>', unsafe_allow_html=True)

# ============================================================
# BOTONES DE ACCIÓN
# ============================================================
def render_botones_accion():
    cols = st.columns([1, 1, 1, 1, 1, 1])
    with cols[0]:
        if st.button("➕ NUEVA", key="btn_nueva", use_container_width=True):
            st.query_params["action"] = "nueva"
            st.rerun()
    with cols[1]:
        if st.button("📥 IMPORT", key="btn_import", use_container_width=True):
            st.query_params["action"] = "importar"
            st.rerun()
    with cols[2]:
        if st.button("📤 EXPORT", key="btn_export", use_container_width=True):
            st.query_params["action"] = "exportar"
            st.rerun()
    
    cols2 = st.columns([1, 1, 1])
    with cols2[0]:
        if st.button("📊 REPORTE", key="btn_reporte", use_container_width=True):
            pass
    with cols2[1]:
        if st.button("🎁 BONUS", key="btn_bonus", use_container_width=True):
            pass
    with cols2[2]:
        if st.button("🏆 QIAF", key="btn_qiaf", use_container_width=True):
            pass

# ============================================================
# BOTONES DE RESTAURANTES
# ============================================================
def render_botones_restaurantes():
    st.markdown("""
    <div style="display: flex; flex-wrap: wrap; gap: 6px; margin: 0.5rem 0;">
        <button class="rest-btn rest-alice">Alice</button>
        <button class="rest-btn rest-arrivals">Arrivals</button>
        <button class="rest-btn rest-lacena">La Cena</button>
        <button class="rest-btn rest-nolimit">NO Limit</button>
        <button class="rest-btn rest-opentable">Open Table</button>
        <button class="rest-btn rest-outlookfw">Outlook-FW</button>
        <button class="rest-btn rest-outlookpc">Outlook-PC</button>
        <button class="rest-btn rest-relaxury">RelaXury</button>
        <button class="rest-btn rest-vtc">VTC</button>
    </div>
    """, unsafe_allow_html=True)

# ============================================================
# FILTROS
# ============================================================
def render_filtros():
    st.markdown('<div class="filter-panel">', unsafe_allow_html=True)
    st.markdown('<div class="filter-label">📅 Date</div>', unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        fecha_sel = st.date_input("", datetime.now().date(), label_visibility="collapsed", key="filtro_fecha")
    with col2:
        if st.button("🔵 APLICAR", use_container_width=True, key="btn_aplicar"):
            st.query_params["fecha_date"] = fecha_sel.strftime("%Y-%m-%d")
            st.query_params["fecha_activa"] = "true"
            st.rerun()
    with col3:
        if st.button("✓ LIMPIAR", use_container_width=True, key="btn_limpiar"):
            st.query_params.clear()
            st.rerun()
    
    if st.button("🔄 RESET", use_container_width=True, key="btn_reset"):
        st.query_params.clear()
        st.rerun()
    
    st.markdown('</div>', unsafe_allow_html=True)

# ============================================================
# BARRA DE BÚSQUEDA
# ============================================================
def render_busqueda():
    busqueda = st.text_input("", placeholder="🔍 Buscar por nombre, teléfono, reserva, VIP, RelaXury...", 
                            label_visibility="collapsed", key="search_input")
    return busqueda

# ============================================================
# TABLA PRINCIPAL CON ESTILO
# ============================================================
def render_tabla(df, busqueda=""):
    if busqueda:
        mask = (
            df["name"].astype(str).str.contains(busqueda, case=False, na=False) |
            df["phone"].astype(str).str.contains(busqueda, case=False, na=False) |
            df["res_number"].astype(str).str.contains(busqueda, case=False, na=False) |
            df["info"].astype(str).str.contains(busqueda, case=False, na=False) |
            df["trans"].astype(str).str.contains(busqueda, case=False, na=False)
        )
        df = df[mask]
    
    if df.empty:
        st.info("No se encontraron reservaciones.")
        return
    
    # Preparar columnas para mostrar
    display_df = df.copy()
    display_df["noches"] = display_df.apply(lambda x: duracion_noches(x["check_in"], x["check_out"]), axis=1)
    
    # Reordenar y renombrar
    cols_mostrar = ["eta", "name", "qty", "room", "check_in", "check_out", "noches", 
                    "res_number", "phone", "email", "info", "ird", "hsk", "rate", "trans"]
    
    display_df = display_df[[c for c in cols_mostrar if c in display_df.columns]]
    
    # Función para colorear celdas
    def colorizar_info(val):
        val_str = str(val).upper()
        if "VIP" in val_str:
            return "color: #FFD700; font-weight: 700;"
        elif "BIRTHDAY" in val_str or "CUMPLE" in val_str:
            return "color: #FF6B6B; font-weight: 700;"
        elif "HONEYMOON" in val_str:
            return "color: #FF69B4; font-weight: 700;"
        elif "TEAM MEMBER" in val_str or "STAFF" in val_str:
            return "color: #FF9800; font-weight: 700;"
        elif "DIAMOND" in val_str:
            return "color: #B9F2FF; font-weight: 700;"
        elif "GOLD" in val_str:
            return "color: #FFD700; font-weight: 700;"
        elif "SILVER" in val_str:
            return "color: #C0C0C0; font-weight: 700;"
        return ""
    
    def colorizar_trans(val):
        val_str = str(val).upper()
        if "RELAXURY" in val_str:
            return "color: #E91E63; font-weight: 700;"
        return ""
    
    # Aplicar estilos
    styled = display_df.style
    if "info" in display_df.columns:
        styled = styled.map(colorizar_info, subset=["info"])
    if "trans" in display_df.columns:
        styled = styled.map(colorizar_trans, subset=["trans"])
    
    # Configurar columnas
    col_config = {
        "eta": st.column_config.TextColumn("ETA", width="small"),
        "name": st.column_config.TextColumn("NAME", width="medium"),
        "qty": st.column_config.TextColumn("QTY", width="small"),
        "room": st.column_config.TextColumn("ROOM", width="small"),
        "check_in": st.column_config.TextColumn("CHECK IN", width="medium"),
        "check_out": st.column_config.TextColumn("CHECK OUT", width="medium"),
        "noches": st.column_config.NumberColumn("🌙", width="small"),
        "res_number": st.column_config.TextColumn("RESERVATION", width="medium"),
        "phone": st.column_config.TextColumn("PHONE", width="medium"),
        "email": st.column_config.TextColumn("EMAIL", width="medium"),
        "info": st.column_config.TextColumn("INFORMATION", width="large"),
        "ird": st.column_config.TextColumn("IRD", width="medium"),
        "hsk": st.column_config.TextColumn("HSK", width="medium"),
        "rate": st.column_config.TextColumn("RATE", width="small"),
        "trans": st.column_config.TextColumn("TRANS", width="medium"),
    }
    
    st.dataframe(
        styled,
        column_config={k: v for k, v in col_config.items() if k in display_df.columns},
        use_container_width=True,
        height=500,
        hide_index=True
    )
    
    st.caption(f"📊 Mostrando {len(display_df)} de {len(df)} reservas")

# ============================================================
# FORMULARIOS
# ============================================================
def formulario_nueva():
    st.markdown("## ➕ Nueva Reservación")
    with st.form("form_nueva"):
        col1, col2 = st.columns(2)
        with col1:
            eta = st.text_input("ETA", value=hora_actual_12h())
            name = st.text_input("Nombre *")
            qty = st.text_input("QTY", value="2")
            room = st.text_input("Room")
            email = st.text_input("Email")
        with col2:
            check_in = st.date_input("Check In", datetime.now().date())
            check_out = st.date_input("Check Out", (datetime.now() + timedelta(days=3)).date())
            res_number = st.text_input("Reservation #")
            phone = st.text_input("Phone")
        
        info = st.text_area("Information (VIP, Birthday, etc.)")
        
        col1, col2 = st.columns(2)
        with col1:
            ird = st.text_input("IRD")
            hsk = st.text_input("HSK")
        with col2:
            rate = st.text_input("Rate")
            trans = st.text_input("Transportation")
        
        submitted = st.form_submit_button("💾 Guardar", use_container_width=True)
        if submitted and name:
            data = {
                "eta": eta, "name": name, "qty": qty, "room": room, "email": email,
                "check_in": check_in.strftime("%B %d, %Y"),
                "check_out": check_out.strftime("%B %d, %Y"),
                "res_number": res_number, "phone": phone, "info": info,
                "ird": ird, "hsk": hsk, "rate": rate, "trans": trans
            }
            insertar_reserva(data)
            st.success("✅ Reservación guardada!")
            time.sleep(1)
            st.query_params.clear()
            st.rerun()
    
    if st.button("← Volver", use_container_width=True):
        st.query_params.clear()
        st.rerun()

def formulario_importar():
    st.markdown("## 📥 Importar Reservaciones")
    archivo = st.file_uploader("Subir archivo Excel o CSV", type=["xlsx", "csv"])
    
    if archivo:
        if archivo.name.endswith(".csv"):
            df_import = pd.read_csv(archivo)
        else:
            df_import = pd.read_excel(archivo)
        
        st.write("Vista previa:")
        st.dataframe(df_import.head())
        
        if st.button("✅ Confirmar Importación", use_container_width=True):
            # Mapear columnas
            columnas_requeridas = ["name", "check_in", "check_out"]
            if not all(c in df_import.columns for c in columnas_requeridas):
                st.error("Faltan columnas requeridas: name, check_in, check_out")
                return
            
            registros = df_import.to_dict("records")
            insertar_batch_reservas(registros)
            st.success(f"✅ {len(registros)} reservaciones importadas!")
            time.sleep(1)
            st.query_params.clear()
            st.rerun()
    
    if st.button("← Volver", use_container_width=True):
        st.query_params.clear()
        st.rerun()

def formulario_exportar(df):
    st.markdown("## 📤 Exportar Reservaciones")
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("📊 Exportar por Categorías (Excel)", use_container_width=True):
            buffer = exportar_excel_por_categorias(df)
            st.download_button(
                label="⬇️ Descargar Excel",
                data=buffer,
                file_name=f"Concierge_Arrivals_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    with col2:
        csv = df.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="⬇️ Descargar CSV",
            data=csv,
            file_name=f"Concierge_Export_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv"
        )
    
    if st.button("← Volver", use_container_width=True):
        st.query_params.clear()
        st.rerun()

# ============================================================
# APP PRINCIPAL
# ============================================================
def main():
    render_header()
    
    df = cargar_reservaciones()
    
    # Barra total
    total = len(df)
    st.markdown(f'<div class="total-bar"><span>📊 TOTAL RESERVAS: {total}</span></div>', unsafe_allow_html=True)
    
    # Routing de páginas
    if mostrar_formulario:
        formulario_nueva()
        return
    elif mostrar_importar:
        formulario_importar()
        return
    elif mostrar_exportar:
        formulario_exportar(df)
        return
    
    # Layout principal
    col_left, col_right = st.columns([1, 4])
    
    with col_left:
        render_checking_out(df)
        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
        render_botones_accion()
        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
        render_botones_restaurantes()
        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
        render_filtros()
    
    with col_right:
        render_categorias(df)
        
        # Filtro de fecha activo
        if fecha_filtro_activo and filtro_fecha_date:
            try:
                fecha_obj = datetime.strptime(filtro_fecha_date, "%Y-%m-%d")
                fecha_str = fecha_obj.strftime("%B %d, %Y")
                df = df[df["check_in"].astype(str).str.contains(fecha_str, na=False)]
                st.info(f"📅 Mostrando check-ins del: {fecha_str}")
            except:
                pass
        
        busqueda = render_busqueda()
        render_tabla(df, busqueda)

if __name__ == "__main__":
    main()
